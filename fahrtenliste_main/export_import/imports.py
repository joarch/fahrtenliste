import datetime
import logging

from django.contrib import messages
from django.utils.safestring import mark_safe
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter

from fahrtenliste_main.export_import.excel import is_empty_value
from fahrtenliste_main.logs import log_entry_change
from fahrtenliste_main.format_util import to_bool

logger = logging.getLogger(__name__)
del logging


def do_import(import_fct, request):
    start = datetime.datetime.now()
    filename = request.FILES['file']
    import_format = request.POST['format']
    dry_run = to_bool(request.POST.get('dry_run', False))

    result = import_fct(request.user, filename, import_format, dry_run=dry_run)
    result_message_success = "Neu: {}, " \
                             "gelöscht: {}, " \
                             "geändert: {}, " \
                             "unverändert: {}, " \
                             "warnung".format(
        len(result["neu"]),
        len(result["geloescht"]),
        len(result["geaendert"]),
        len(result["unveraendert"]),
        len(result["warnung"]))
    result_message = "<a href='#neu'>Neu</a> {}, " \
                     "<a href='#geloescht'>gelöscht</a> {}, " \
                     "<a href='#geaendert'>geändert</a> {}, " \
                     "<a href='#unveraendert'>unverändert</a> {}," \
                     "<a href='#warnungen'>warnung</a> {}".format(
        len(result["neu"]),
        len(result["geloescht"]),
        len(result["geaendert"]),
        len(result["unveraendert"]),
        len(result["warnung"]))
    end = datetime.datetime.now()
    result["import_dauer"] = str_duration(end - start)
    result["result"] = result_message
    messages.success(request, mark_safe("Fertig: " + result_message))
    log_entry_change(request.user, "adresse", "Adressen Import",
                     result["filename"] + " (" + result["format"] + "): " + result_message_success)
    return result


def str_duration(duration):
    mm, ss = divmod(duration.seconds, 60)
    hh, mm = divmod(mm, 60)
    s = "%d:%02d:%02d" % (hh, mm, ss)
    if duration.days:
        def plural(n):
            return n, abs(n) != 1 and "s" or ""

        s = ("%d day%s, " % plural(duration.days)) + s
    return s


def import_von_excel(file_path, input_format):
    mehrere_sheets = input_format.get("mehrere_sheets", False)
    start_value = input_format.get("start_value")
    start_value_column = input_format.get("start_value_column")
    columns = input_format["columns"]
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)

    # sheets ermitteln
    wss = []
    if not mehrere_sheets:
        wss.append(wb[wb.get_sheet_names()[0]])
    else:
        for index, ws_name in enumerate(wb.get_sheet_names()):
            wss.append(wb[ws_name])

    result = []
    for ws in wss:
        start_row = input_format["start_row"]
        logger.verbose("Import {}".format(ws.title))
        skip = True if start_value is not None else False
        for idx_row, row in enumerate(ws.rows):
            logger.verbose("Excel Import {}".format(idx_row + 1))
            row_number = idx_row + 1

            if skip:
                # skip solange bis start Wert aus Start Column gefunden wurde
                value = ws.cell(column=column_index_from_string(start_value_column), row=row_number).value
                if value != start_value:
                    if row_number > 100:
                        print("Abbruch leere Tabelle")
                        # Abbruch kein Start Wert gefunden
                        # z.B. leerer sheet
                        break
                    else:
                        continue
                else:
                    # Start Wert gefunden
                    # ab jetzt bis Start Row weitergehen
                    start_row = row + start_row - 1
                    skip = False

            if row_number < start_row:
                continue

            destination_dict = dict()
            empty_row = True

            cell_idx = 1
            for cell in row:
                column_letter = get_column_letter(cell_idx)
                attribute = columns.get(column_letter)
                if attribute is not None:
                    if attribute[0:1] == "#":
                        # Kommentar ignorieren
                        continue
                    value = cell.value
                    empty_row = empty_row and is_empty_value(value)
                    destination_dict[attribute] = value
                cell_idx += 1
            if empty_row:
                break

            """
            # zu langsam!
            for column, attribute in columns.items():
                if attribute[0:1] == "#":
                    # Kommentar ignorieren
                    continue
                value = ws.cell(column=column_index_from_string(column), row=row_number).value
                empty_row = empty_row and is_empty_value(value)
                destination_dict[attribute] = value
            if empty_row:
                break
            """

            result.append(destination_dict)

    wb.close()

    return result

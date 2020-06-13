import datetime
import logging
import os
import re
import tempfile
from collections import namedtuple
from decimal import Decimal

from django.utils.encoding import smart_str
from django.views.static import serve
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.fonts import Font
from openpyxl.utils.cell import get_column_letter

logger = logging.getLogger(__name__)
del logging


def serve_export(request, file_path):
    file_name = os.path.basename(file_path)
    dir_name = os.path.dirname(file_path)
    response = serve(request, file_name, dir_name)
    response['Content-Disposition'] = 'attachment; filename=%s' % smart_str(file_name)
    response['X-Sendfile'] = smart_str(file_name)
    return response


def export_nach_excel(namedtuple_in: namedtuple, detail_name: str, data_list: list, filename_postfix="",
                      filename_in=None, mappings=dict(), export_name=None, export_beschreibung=None):
    if filename_in is None:
        if export_name is None:
            # Export Name aus Nametuple ableiten
            export_name = namedtuple_in.__name__.capitalize()
        export_beschreibung = None
    else:
        if export_name is not None or export_beschreibung is not None:
            raise RuntimeError("'filename_in' und 'export_name' oder 'export_beschreibung' "
                               "dürfen nicht zusammen angegeben werden!")
        # filename_in wird verwendet um
        # 1. den ersten Teil bis zum " " als echten Dateinamen zu verwenden
        # 2. eine zusätzliche Infozeile auszugeben
        # Hinweis: bei zu langen Dateinamen unter Windows/Excel, kommt es zu einer kaputten Excel Datei,
        # die beim Öffnen dann immer zuerst autom. durch Excel repariert werden muss
        if "~" in filename_in:
            # Filename~Export Name~Beschreibung
            export_name = filename_in.split("~")[0]
            export_beschreibung = filename_in.split("~")[1] + ": " + filename_in.split("~")[2]
        else:
            export_name = filename_in if " " not in filename_in else filename_in.split(" ")[0]
            export_beschreibung = filename_in

    timestamp = datetime.datetime.now().strftime("_Export_%Y-%m-%d_%H-%M")
    filename = "{}_{}_{}_{}.xlsx".format(filename_postfix, export_name, detail_name, timestamp)
    filename = re.sub(r'^_', '', filename)
    filename = filename.replace(" ", "_")
    filename = re.sub(r'_{2,9}', '_', filename)

    temp_dir = tempfile.gettempdir()
    export_file_path = os.path.join(temp_dir, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = export_name

    column_header = PatternFill(start_color='AAAAAA',
                                end_color='AAAAAA',
                                fill_type='solid')

    if export_beschreibung is not None:
        row = 4
        ws.cell(column=1, row=1, value=export_beschreibung)
        ws.cell(column=1, row=2, value="Datum: {}".format(datetime.datetime.now().strftime("%d.%m.%Y %H:%M")))
        ws.merge_cells('A1:H1')
        ws.merge_cells('A2:H2')
    else:
        row = 1

    column = 1
    for name in namedtuple_in._fields:
        ws.cell(column=column, row=row, value=name.capitalize())
        col = ws["{}{}".format(get_column_letter(column), row)]
        col.font = Font(bold=True)
        col.fill = column_header
        column += 1

    row += 1
    for data in data_list:
        column = 1
        for name, value in data._asdict().items():
            try:
                if name in mappings:
                    mapped_value = mappings[name](value)
                else:
                    mapped_value = map_str_value(value)
            except Exception:
                logger.warning("Fehler bei mapped_value, Wert: '{}'".format(value))
                mapped_value = value

            cell = ws.cell(column=column, row=row, value=mapped_value)
            if isinstance(mapped_value, datetime.date):
                cell.number_format = 'DD.MM.YYYY'
            column += 1
        row += 1

    wb.save(filename=export_file_path)
    return export_file_path


def map_number(value):
    if value is None:
        return None
    return Decimal(value.replace('.', '').replace(' ', '').replace(',', '.'))


def map_eur(value):
    return map_number(value.replace("€", "").replace("EUR", ""))


def map_datum(value):
    return datetime.datetime.strptime(value, "%d.%m.%Y").date()


def map_str_value(value):
    if type(value) != str:
        return value

    if value == "undefined" or value.startswith("Sum: "):
        # clean up
        value = ""

    if value == "true":
        value = "ja"

    if value == "false":
        value = "nein"

    if "€" in value \
            and (value[0].isnumeric() or len(value) >= 2 and value[0] == "-" and value[1].isnumeric()):
        return map_eur(value)

    eur_pattern = re.compile("^-?[0-9]+,?[0-9]* (€|EUR)$")
    if eur_pattern.match(value):
        return map_eur(value)

    datum_pattern = re.compile("^[0-9]{2}\.[0-9]{2}\.[0-9]{4}$")
    if datum_pattern.match(value):
        return map_datum(value)

    decimal_pattern = re.compile("^-?[0-9]+,?[0-9]*$")
    if decimal_pattern.match(value):
        return map_number(value)

    return value


def none_mapping(value):
    return value
